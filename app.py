from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
import re
from threading import Timer
import os
from typing import Any
import sys
import webbrowser
import json

import pandas as pd
from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook

BASE_DIR = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
app = Flask(__name__, template_folder=str(BASE_DIR / "templates"))


@dataclass
class ColumnMap:
    order_number: int = 4  # E
    order_date: int = 3  # D
    order_customer: int = 6  # G
    order_part: int = 14  # O
    order_net_sales: int = 20  # U

    quote_number: int = 0  # A
    quote_line: int = 1  # B
    quote_part: int = 2  # C
    quote_customer: int = 35  # AJ
    quote_customer_name: int = 36  # AK
    quote_date: int = 48  # AW
    quote_user: int = 61  # BJ
    quote_ext_price: int = 13  # N


COLUMNS = ColumnMap()


def _safe_get(df: pd.DataFrame, idx: int, name: str) -> pd.Series:
    if idx >= df.shape[1]:
        raise ValueError(
            f"Expected column {name} at index {idx} (Excel letter position), "
            f"but file only has {df.shape[1]} columns."
        )
    return df.iloc[:, idx]


def load_orders(file_obj: Any) -> pd.DataFrame:
    raw = pd.read_excel(file_obj)
    orders = pd.DataFrame(
        {
            "order_number": _safe_get(raw, COLUMNS.order_number, "Order Number (E)").astype(str).str.strip(),
            "order_date": pd.to_datetime(_safe_get(raw, COLUMNS.order_date, "Order Date (D)"), errors="coerce"),
            "customer_id": _safe_get(raw, COLUMNS.order_customer, "Customer ID (G)").astype(str).str.strip(),
            "part_number": _safe_get(raw, COLUMNS.order_part, "Part Number (O)").astype(str).str.strip().str.upper(),
            "net_sales": pd.to_numeric(_safe_get(raw, COLUMNS.order_net_sales, "Net Sales (U)"), errors="coerce").fillna(0.0),
        }
    )
    orders = orders.dropna(subset=["order_date"]).copy()
    return orders[(orders["part_number"].ne("")) & (orders["customer_id"].ne(""))]


def load_quotes(file_obj: Any) -> pd.DataFrame:
    raw = pd.read_excel(file_obj)
    quotes = pd.DataFrame(
        {
            "quote_number": _safe_get(raw, COLUMNS.quote_number, "Quote Number (A)").astype(str).str.strip(),
            "line_number": _safe_get(raw, COLUMNS.quote_line, "Line Number (B)"),
            "part_number": _safe_get(raw, COLUMNS.quote_part, "Part Number (C)").astype(str).str.strip().str.upper(),
            "customer_id": _safe_get(raw, COLUMNS.quote_customer, "Customer ID (AJ)").astype(str).str.strip(),
            "customer_name": _safe_get(raw, COLUMNS.quote_customer_name, "Customer Name (AK)").astype(str).str.strip(),
            "quote_date": pd.to_datetime(_safe_get(raw, COLUMNS.quote_date, "Date Quoted (AW)"), errors="coerce"),
            "user_id": _safe_get(raw, COLUMNS.quote_user, "User ID (BJ)").astype(str).str.strip(),
            "ext_price": pd.to_numeric(_safe_get(raw, COLUMNS.quote_ext_price, "Ext. Price (N)"), errors="coerce").fillna(0.0),
        }
    )
    quotes = quotes.dropna(subset=["quote_date"]).copy()
    return quotes[(quotes["part_number"].ne("")) & (quotes["quote_number"].ne(""))]


def run_conversion(orders: pd.DataFrame, quotes: pd.DataFrame, min_line_match_ratio: float = 0.90) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if quotes.empty:
        empty = quotes.assign(converted=False)
        return empty, pd.DataFrame(), pd.DataFrame()

    merged = quotes.merge(
        orders,
        on=["customer_id", "part_number"],
        how="left",
        suffixes=("_quote", "_order"),
    )

    merged["days_to_convert"] = (merged["order_date"] - merged["quote_date"]).dt.days
    merged["valid_conversion"] = merged["days_to_convert"].notna() & (merged["days_to_convert"] >= 0)

    valid = merged[merged["valid_conversion"]].copy()

    quote_lines = (
        quotes.groupby("quote_number", dropna=False)
        .agg(total_lines=("line_number", "nunique"))
        .reset_index()
    )

    if valid.empty:
        line_output = quotes.copy()
        line_output["order_number"] = pd.NA
        line_output["order_date"] = pd.NaT
        line_output["net_sales"] = 0.0
        line_output["days_to_convert"] = pd.NA
        line_output["converted"] = False

        def _first_non_empty(series: pd.Series) -> str:
            for value in series:
                if pd.notna(value) and str(value).strip() != "":
                    return str(value).strip()
            return ""

        quote_output = (
            line_output.groupby(["quote_number"], dropna=False)
            .agg(
                customer_id=("customer_id", _first_non_empty),
                customer_name=("customer_name", _first_non_empty),
                user_id=("user_id", _first_non_empty),
                quote_date=("quote_date", "min"),
                parts_quoted=("part_number", "nunique"),
                total_lines=("line_number", "nunique"),
                quote_amount=("ext_price", "sum"),
            )
            .reset_index()
        )
        quote_output["matched_orders"] = ""
        quote_output["converted_net_sales"] = 0.0
        quote_output["converted_lines"] = 0
        quote_output["line_match_rate"] = 0.0
        quote_output["converted"] = False
        quote_output["follow_up_needed"] = True
    else:
        line_order_best = (
            valid.sort_values("days_to_convert")
            .groupby(["quote_number", "line_number", "order_number"], as_index=False)
            .first()
        )

        order_coverage = (
            line_order_best.groupby(["quote_number", "order_number"], as_index=False)
            .agg(
                matched_lines=("line_number", "nunique"),
                order_date=("order_date", "min"),
                converted_net_sales=("net_sales", "sum"),
            )
            .merge(quote_lines, on="quote_number", how="left")
        )
        order_coverage["line_match_rate"] = (order_coverage["matched_lines"] / order_coverage["total_lines"]).fillna(0)

        best_order = (
            order_coverage.sort_values(["quote_number", "line_match_rate", "matched_lines", "order_date"], ascending=[True, False, False, True])
            .groupby("quote_number", as_index=False)
            .first()
        )
        best_order["converted"] = best_order["line_match_rate"] >= min_line_match_ratio

        winning_line_matches = line_order_best.merge(
            best_order[["quote_number", "order_number", "converted"]],
            on=["quote_number", "order_number"],
            how="inner",
        )
        winning_line_matches = winning_line_matches[winning_line_matches["converted"]].copy()

        converted_lines = (
            winning_line_matches.groupby("quote_number", as_index=False)
            .agg(converted_lines=("line_number", "nunique"))
        )

        line_output = quotes.merge(
            winning_line_matches[["quote_number", "line_number", "order_number", "order_date", "net_sales", "days_to_convert"]],
            on=["quote_number", "line_number"],
            how="left",
        )
        line_output["converted"] = line_output["order_number"].notna()

        def _first_non_empty(series: pd.Series) -> str:
            for value in series:
                if pd.notna(value) and str(value).strip() != "":
                    return str(value).strip()
            return ""

        quote_output = (
            line_output.groupby(["quote_number"], dropna=False)
            .agg(
                customer_id=("customer_id", _first_non_empty),
                customer_name=("customer_name", _first_non_empty),
                user_id=("user_id", _first_non_empty),
                quote_date=("quote_date", "min"),
                parts_quoted=("part_number", "nunique"),
                total_lines=("line_number", "nunique"),
                quote_amount=("ext_price", "sum"),
            )
            .reset_index()
            .merge(best_order[["quote_number", "order_number", "order_date", "converted_net_sales", "line_match_rate", "converted"]], on="quote_number", how="left")
            .merge(converted_lines, on="quote_number", how="left")
        )
        quote_output["converted_lines"] = quote_output["converted_lines"].fillna(0).astype(int)
        quote_output["converted"] = quote_output["converted"].astype("boolean").fillna(False).astype(bool)
        quote_output["line_match_rate"] = quote_output["line_match_rate"].fillna(0.0)
        quote_output["converted_net_sales"] = quote_output["converted_net_sales"].fillna(0.0)
        quote_output["matched_orders"] = quote_output.apply(
            lambda r: str(r["order_number"]) if bool(r["converted"]) and pd.notna(r["order_number"]) else "",
            axis=1,
        )
        quote_output = quote_output.drop(columns=["order_number"])
        quote_output["follow_up_needed"] = ~quote_output["converted"]

    rep_summary = (
        quote_output.groupby("user_id", dropna=False)
        .agg(
            consolidated_quotes=("quote_number", "count"),
            converted_quotes=("converted", "sum"),
            converted_net_sales=("converted_net_sales", "sum"),
        )
        .reset_index()
    )
    rep_summary["conversion_rate"] = (rep_summary["converted_quotes"] / rep_summary["consolidated_quotes"]).fillna(0)

    return line_output, rep_summary, quote_output


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace("_", " ")


def _find_header_row(ws) -> int:
    for row_num in range(1, min(ws.max_row, 30) + 1):
        labels = [_normalize_header(c.value) for c in ws[row_num] if c.value is not None]
        if any("quote" in v for v in labels) and (any("customer" in v for v in labels) or any("date" in v for v in labels)):
            return row_num
    return 1


def _column_map_from_header(ws, header_row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        label = _normalize_header(ws.cell(row=header_row, column=col).value)
        if not label:
            continue
        if (("quote" in label and "#" in label) or "quote number" in label or label == "quote"):
            mapping.setdefault("quote_number", col)
        elif "customer" in label or "cust" in label:
            mapping.setdefault("customer_id", col)
        elif "date" in label:
            mapping.setdefault("quote_date", col)
        elif "rep" in label or "user" in label or "owner" in label or "assignee" in label:
            mapping.setdefault("user_id", col)
        elif "quote amount" in label or ("amount" in label and "order" not in label):
            mapping.setdefault("quote_amount", col)
        elif "part" in label and ("count" in label or "qty" in label or "quoted" in label):
            mapping.setdefault("parts_quoted", col)
        elif "won/lost" in label or ("won" in label and "lost" in label):
            mapping.setdefault("follow_up_needed", col)
        elif "actual order" in label:
            mapping.setdefault("converted_net_sales", col)
        elif "order" in label:
            mapping.setdefault("matched_orders", col)
        elif "net" in label or "sales" in label:
            mapping.setdefault("converted_net_sales", col)
        elif "status" in label or "follow" in label:
            mapping.setdefault("follow_up_needed", col)
        elif "note" in label:
            mapping.setdefault("notes", col)
    return mapping


def _clear_sheet_data(ws, start_row: int) -> None:
    if ws.max_row <= start_row:
        return

    first_data_row = start_row + 1
    last_data_row = ws.max_row

    if ws._tables:
        table = next(iter(ws._tables.values()))
        _, last_cell = table.ref.split(":")
        last_data_row = ws[last_cell].row

    for row in range(first_data_row, last_data_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.data_type == "f":
                continue
            cell.value = None


def _normalize_person(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (value or "").lower())


def _display_name_from_sheet(sheet_title: str) -> str:
    return sheet_title.split("(")[0].strip()


def _assign_sheet_for_rep(workbook, rep: str):
    non_summary = [ws for ws in workbook.worksheets if "summary" not in ws.title.lower()]

    rep_clean = _normalize_person(rep)
    if not rep_clean:
        return non_summary[0] if non_summary else workbook.worksheets[0]

    for ws in non_summary:
        if _normalize_person(ws.title) == rep_clean:
            return ws

    for ws in non_summary:
        ws_clean = _normalize_person(ws.title)
        if rep_clean in ws_clean or ws_clean in rep_clean:
            return ws

    # Match user IDs like bschrader to sheet names like Brent Schrader by last name.
    for ws in non_summary:
        ws_parts = [p for p in re.split(r"\s+", ws.title.strip()) if p]
        if ws_parts:
            last_name = _normalize_person(ws_parts[-1])
            if last_name and rep_clean.endswith(last_name):
                return ws

    # Master Allocation should be fallback only when no named tab could be identified.
    for ws in non_summary:
        if "master" not in ws.title.lower():
            return ws

    for ws in non_summary:
        if rep_clean and rep_clean in _normalize_person(ws.title):
            return ws
    return non_summary[0] if non_summary else workbook.worksheets[0]


def build_follow_up_workbook(template_bytes: bytes, quotes_for_followup: pd.DataFrame) -> bytes:
    wb = load_workbook(BytesIO(template_bytes))

    quotes_for_followup = quotes_for_followup.copy()
    if quotes_for_followup.empty:
        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    target_sheets = {row.user_id: _assign_sheet_for_rep(wb, row.user_id) for row in quotes_for_followup[["user_id"]].drop_duplicates().itertuples()}

    for ws in set(target_sheets.values()):
        header_row = _find_header_row(ws)
        _clear_sheet_data(ws, header_row)

    for rep, rep_df in quotes_for_followup.groupby("user_id", dropna=False):
        ws = target_sheets.get(rep)
        if ws is None:
            continue

        header_row = _find_header_row(ws)
        col_map = _column_map_from_header(ws, header_row)

        if not col_map:
            continue

        row_num = header_row + 1

        table_last_row = ws.max_row
        if ws._tables:
            table = next(iter(ws._tables.values()))
            _, last_cell = table.ref.split(":")
            table_last_row = ws[last_cell].row

        for rec in rep_df.sort_values(["quote_date", "quote_number"]).to_dict("records"):
            if row_num > table_last_row:
                break

            assignee_name = _display_name_from_sheet(ws.title)
            if "quote_number" in col_map:
                ws.cell(row=row_num, column=col_map["quote_number"], value=rec.get("quote_number"))
            if "customer_id" in col_map:
                ws.cell(
                    row=row_num,
                    column=col_map["customer_id"],
                    value=rec.get("customer_name") or rec.get("customer_id"),
                )
            if "quote_date" in col_map:
                ws.cell(row=row_num, column=col_map["quote_date"], value=rec.get("quote_date"))
            if "user_id" in col_map:
                ws.cell(row=row_num, column=col_map["user_id"], value=assignee_name)
            if "quote_amount" in col_map:
                quote_amount = rec.get("quote_amount")
                ws.cell(row=row_num, column=col_map["quote_amount"], value=float(quote_amount) if pd.notna(quote_amount) else None)
            if "parts_quoted" in col_map:
                ws.cell(row=row_num, column=col_map["parts_quoted"], value=int(rec.get("parts_quoted") or 0))
            if "converted_net_sales" in col_map:
                ws.cell(row=row_num, column=col_map["converted_net_sales"], value=float(rec.get("converted_net_sales") or 0))
            if "follow_up_needed" in col_map:
                ws.cell(
                    row=row_num,
                    column=col_map["follow_up_needed"],
                    value="",
                )
            if "matched_orders" in col_map:
                ws.cell(row=row_num, column=col_map["matched_orders"], value=rec.get("matched_orders"))
            if "notes" in col_map:
                ws.cell(row=row_num, column=col_map["notes"], value="")
            row_num += 1

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def _resolve_template_bytes(uploaded_template) -> bytes:
    if uploaded_template and uploaded_template.filename:
        return uploaded_template.read()

    local_template = BASE_DIR / "assets" / "Parts Follow Up Template.xlsx"
    if local_template.exists():
        return local_template.read_bytes()

    raise ValueError(
        "Template file is required. Upload 'Parts Follow Up Template.xlsx' in the form, "
        "or place it in assets/Parts Follow Up Template.xlsx."
    )



def _format_currency(value: Any) -> str:
    if pd.isna(value):
        return ""
    return f"${float(value):,.2f}"


def _format_percent(value: Any) -> str:
    if pd.isna(value):
        return ""
    return f"{float(value) * 100:.2f}%"


@app.route("/", methods=["GET", "POST"])
def index():
    quote_results_html = None
    rep_results_html = None
    rep_chart_data = "[]"
    followup_chart_data = "[]"
    error = None

    if request.method == "POST":
        try:
            order_file = request.files.get("order_file")
            quote_file = request.files.get("quote_file")
            template_file = request.files.get("template_file")
            min_quote_amount = float(request.form.get("min_quote_amount") or 2000)

            if not order_file or not quote_file:
                raise ValueError("Please upload both an order log file and a quote line summary file.")

            template_bytes = _resolve_template_bytes(template_file)
            orders = load_orders(order_file)
            quotes = load_quotes(quote_file)

            line_results, rep_summary, quote_results = run_conversion(orders, quotes, min_line_match_ratio=0.90)

            min_amount_mask = quote_results["quote_amount"].fillna(0) >= min_quote_amount
            follow_up_quotes = quote_results[quote_results["follow_up_needed"] & min_amount_mask].copy()
            generated_report = build_follow_up_workbook(template_bytes, follow_up_quotes)

            quote_results_display = quote_results.sort_values(["quote_date", "quote_number"]).copy()
            quote_results_display["converted_net_sales"] = quote_results_display["converted_net_sales"].map(_format_currency)
            quote_results_display["quote_amount"] = quote_results_display["quote_amount"].map(_format_currency)
            quote_results_display["line_match_rate"] = quote_results_display["line_match_rate"].map(_format_percent)
            quote_results_display = quote_results_display.rename(
                columns={
                    "quote_number": "Quote Number",
                    "customer_id": "Customer ID",
                    "customer_name": "Customer Name",
                    "user_id": "Parts Rep",
                    "quote_date": "Quote Date",
                    "parts_quoted": "Parts Quoted",
                    "matched_orders": "Matched Orders",
                    "converted_net_sales": "Converted Net Sales",
                    "converted_lines": "Converted Lines",
                    "total_lines": "Total Lines",
                    "line_match_rate": "Line Match Rate",
                    "converted": "Converted",
                    "follow_up_needed": "Follow Up Needed",
                    "quote_amount": "Quote Amount",
                }
            )
            quote_results_html = quote_results_display.to_html(
                index=False,
                classes="results-table sortable-table",
                table_id="quote-results-table",
            )

            rep_results_display = rep_summary.sort_values("conversion_rate", ascending=False).copy()
            rep_results_display["converted_net_sales"] = rep_results_display["converted_net_sales"].map(_format_currency)
            rep_results_display["conversion_rate"] = rep_results_display["conversion_rate"].map(_format_percent)
            rep_results_display = rep_results_display.rename(
                columns={
                    "user_id": "Parts Rep",
                    "consolidated_quotes": "Number of Quotes",
                    "converted_quotes": "Number of Converted Quotes",
                    "converted_net_sales": "Converted Net Sales",
                    "conversion_rate": "Conversion Rate",
                }
            )
            rep_results_html = rep_results_display.to_html(
                index=False,
                classes="results-table sortable-table",
                table_id="rep-summary-table",
            )

            rep_chart = rep_summary.sort_values("conversion_rate", ascending=False).copy()
            rep_chart_data = json.dumps(
                [
                    {
                        "rep": str(r.user_id or "Unassigned"),
                        "conversion_rate": float(r.conversion_rate),
                    }
                    for r in rep_chart.itertuples()
                ]
            )
            followup_counts = (
                quote_results[quote_results["follow_up_needed"]]
                .groupby("user_id", dropna=False)
                .size()
                .reset_index(name="count")
                .sort_values("count", ascending=False)
            )
            followup_chart_data = json.dumps(
                [
                    {
                        "rep": str(r.user_id or "Unassigned"),
                        "count": int(r.count),
                    }
                    for r in followup_counts.itertuples()
                ]
            )

            app.config["last_quote_results"] = quote_results
            app.config["last_rep_results"] = rep_summary
            app.config["last_followup_workbook"] = generated_report
        except Exception as exc:
            error = str(exc)

    return render_template(
        "index.html",
        quote_results_html=quote_results_html,
        rep_results_html=rep_results_html,
        rep_chart_data=rep_chart_data,
        followup_chart_data=followup_chart_data,
        error=error,
    )


@app.route("/download/followup")
def download_followup():
    workbook_bytes = app.config.get("last_followup_workbook")
    if not workbook_bytes:
        return "No report available yet. Run an analysis first.", 400

    return send_file(
        BytesIO(workbook_bytes),
        as_attachment=True,
        download_name="Parts_Follow_Up_Output.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/favicon.ico")
def favicon():
    return send_file(BASE_DIR / "assets" / "app.ico", mimetype="image/x-icon")


def _open_browser() -> None:
    webbrowser.open("http://127.0.0.1:8000")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "8000"))
    is_local_dev = not os.environ.get("CI") and not os.environ.get("PORT")
    if is_local_dev:
        Timer(1.0, _open_browser).start()
    app.run(host="0.0.0.0", port=port, debug=False, use_reloader=False)
